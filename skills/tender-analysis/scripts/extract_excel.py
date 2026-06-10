#!/usr/bin/env python3
"""
提取 Excel 文档内容
支持 .xls 和 .xlsx 格式
"""
import sys
import os

def extract_excel(file_path):
    """提取 Excel 文档内容"""
    try:
        # 尝试使用 openpyxl（.xlsx）
        if file_path.lower().endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                content = []
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    content.append(f"\n=== 工作表: {sheet_name} ===\n")
                    
                    for row in sheet.iter_rows(values_only=True):
                        row_text = [str(cell) if cell is not None else '' for cell in row]
                        if any(row_text):
                            content.append(" | ".join(row_text))
                
                return "\n".join(content)
            except ImportError:
                print("错误: 需要安装 openpyxl: pip install openpyxl", file=sys.stderr)
                return None
            except Exception as e:
                print(f"openpyxl 处理失败: {e}", file=sys.stderr)
        
        # 尝试使用 xlrd（.xls）
        elif file_path.lower().endswith('.xls'):
            try:
                import xlrd
                workbook = xlrd.open_workbook(file_path)
                content = []
                
                for sheet_name in workbook.sheet_names():
                    sheet = workbook.sheet_by_name(sheet_name)
                    content.append(f"\n=== 工作表: {sheet_name} ===\n")
                    
                    for row_idx in range(sheet.nrows):
                        row = sheet.row_values(row_idx)
                        row_text = [str(cell) if cell else '' for cell in row]
                        if any(row_text):
                            content.append(" | ".join(row_text))
                
                return "\n".join(content)
            except ImportError:
                print("错误: 需要安装 xlrd: pip install xlrd", file=sys.stderr)
                return None
            except Exception as e:
                print(f"xlrd 处理失败: {e}", file=sys.stderr)
        
        return None
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        return None

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python extract_excel.py <文件路径>")
        print("支持格式: .xls, .xlsx")
        sys.exit(1)
    
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"错误: 文件不存在: {file_path}", file=sys.stderr)
        sys.exit(1)
    
    content = extract_excel(file_path)
    
    if content:
        print(content)
    else:
        sys.exit(1)
